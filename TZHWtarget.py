# _*_ coding:utf-8 _*_
# @Time : 2023/6/3 20:57
# @Author: hu4x1n


import requests
import openpyxl
import time

# 修改 1
header = {
    "Host": "nsiscp.jscert.cn:8078",
    "Connection": "close",
    "sec-ch-ua": "Not?A_Brand;v=8, Chromium;v=108, Google Chrome;v=108",
    "Authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwOlwvXC8xMjcuMC4wLjE6MTAwMDFcL2FwaVwvYXV0aFwvbG9naW4iLCJpYXQiOjE3NDA1NTY0NjksImV4cCI6MTc0MDY0Mjg2OSwibmJmIjoxNzQwNTU2NDY5LCJqdGkiOiIyZ05RR2JGZGtyd0tzYXlNIiwic3ViIjoiRkU0QUNGNzUzMUY5MDRDQjZGOUE1OTIzNjIyODAwMjEiLCJwcnYiOiIyM2JkNWM4OTQ5ZjYwMGFkYjM5ZTcwMWM0MDA4NzJkYjdhNTk3NmY3In0.um0xS0Omh3YQELxHnHDNzTAmqU4PcRPqo5_XC2vThrI",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36",
    "sec-ch-ua-platform": "Windows",
    "Cookie": "assetreport_session=eyJpdiI6IjZ1bmRYVzgxNjJxdU5BR0E1QVZNSlE9PSIsInZhbHVlIjoic1dJY2VCNmErNUFMYkhSZzhqZU8xNC9iOU5SNE5yWGVHd1Uvdm1Rcm1WOWF5NXkyY3FOcUhQTCtHU3RQTU9DRnd1QkRZdnp6a25oUXdQU3RnblZSdEZ3RElXbDJpbkJYbUNkMTIxNk93Zm92R1Q0ZTIvb1ZMdUhac3FmSmVScU0iLCJtYWMiOiJmMWJkNDgxMmFkYTU0OThlNDAwODBmNDIwNjdhMDM4ZjIzMThhZGIwZTM0NmJkOTg1NGMyMGM2M2VlZTI3MDQxIiwidGFnIjoiIn0%3D",
}
page = 1
# 修改 2
url = "https://nsiscp.jscert.cn:8078/api/activity/648E938218A5900BEA5667A7047113D1/asset/white_list"

# 在当前文件目录下创建一个excel文件，文件名为TZHWdata.xlsx
workbook = openpyxl.load_workbook("TZHWdata.xlsx")
worksheet = workbook["Sheet1"]
worksheet["A1"] = "assetUrl"
worksheet["B1"] = "realIp"
worksheet["C1"] = "eventNum"
worksheet["D1"] = "noCheckEventNum"
worksheet["E1"] = "findType"
worksheet["F1"] = "createdAt"
worksheet["G1"] = "ingorevulNum"
worksheet["H1"] = "submitRestrict"
worksheet["I1"] = "riskTypeId"
worksheet["J1"] = "riskTypeName"
row = 2
for n in range(1, 200):
    postdata = {
        "page": n,
        "pageSize": "50",
        "assetUrl": "",
        "findType": "",
        "startTime": "",
        "endTime": "",
    }
    response = requests.post(url=url, headers=header, data=postdata)

    data_list = response.json()["data"]["list"]
    for dic in data_list:
        assetUrl = dic["assetUrl"]
        realIp = dic["realIp"]
        eventNum = dic["eventNum"]
        noCheckEventNum = dic["noCheckEventNum"]
        findType = dic["findType"]
        createdAt = dic["createdAt"]
        ingorevulNum = dic["ingorevulNum"]
        submitRestrict = dic["submitRestrict"]
        riskTypeId = dic["riskTypeId"]
        riskTypeName = dic["riskTypeName"]
        print(f"assetUrl:{assetUrl},createdAt:{createdAt}")

        worksheet.cell(row=row, column=1).value = assetUrl
        worksheet.cell(row=row, column=2).value = realIp
        worksheet.cell(row=row, column=3).value = eventNum
        worksheet.cell(row=row, column=4).value = noCheckEventNum
        worksheet.cell(row=row, column=5).value = findType
        worksheet.cell(row=row, column=6).value = createdAt
        worksheet.cell(row=row, column=7).value = ingorevulNum
        worksheet.cell(row=row, column=8).value = submitRestrict
        worksheet.cell(row=row, column=9).value = riskTypeId
        worksheet.cell(row=row, column=10).value = riskTypeName
        row = row + 1
        time.sleep(0.1)

    print(f"第{n}页爬取完成")
    # 修改保存路径
    workbook.save("TZHWdata.xlsx")

workbook.close()
